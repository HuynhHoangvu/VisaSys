#!/usr/bin/env python3
"""
Usage:
  python3 gen_salary.py <input.json> <output.pdf>           # phieu luong ca nhan (PDF)
  python3 gen_salary.py summary <input.json> <output.pdf>   # bang luong tong (PDF)
  python3 gen_salary.py excel <input.json> <output.xlsx>    # bang luong tong (Excel)
  python3 gen_salary.py slips <input.json> <output.xlsx>    # phieu luong ca nhan (Excel)
"""
import sys
import json
import os
import math
from reportlab.lib.pagesizes import A4, landscape, A3
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

FONT = 'Helvetica'
FONT_BOLD = 'Helvetica-Bold'

_font_candidates = [
    ('C:/Windows/Fonts/arial.ttf',   'C:/Windows/Fonts/arialbd.ttf'),
    ('/usr/share/fonts/dejavu/DejaVuSans.ttf',        '/usr/share/fonts/dejavu/DejaVuSans-Bold.ttf'),
    ('/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf', '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf'),
]
for _reg, _bold in _font_candidates:
    if os.path.exists(_reg) and os.path.exists(_bold):
        try:
            pdfmetrics.registerFont(TTFont('UniFont',      _reg))
            pdfmetrics.registerFont(TTFont('UniFont-Bold', _bold))
            FONT      = 'UniFont'
            FONT_BOLD = 'UniFont-Bold'
            break
        except Exception:
            pass

def fmt(amount):
    if not amount: return "0"
    try: return f"{int(amount):,}".replace(",", ".")
    except: return str(amount)

def money_amount(v):
    """JSON có thể trả chuỗi — Excel SUM() bỏ qua ô text nên cột H/J sai; ép float."""
    try:
        if v is None:
            return 0.0
        if isinstance(v, bool):
            return 0.0
        if isinstance(v, str):
            s = v.strip().replace("\u00a0", "").replace(" ", "").replace(",", "")
            if not s:
                return 0.0
            return float(s)
        x = float(v)
        if math.isnan(x) or math.isinf(x):
            return 0.0
        return x
    except (TypeError, ValueError):
        return 0.0

def p(text, font=None, size=9, align=0, color=colors.black):
    f = font or FONT
    return Paragraph(str(text), ParagraphStyle('x', fontName=f, fontSize=size, alignment=align, textColor=color, leading=size + 3))

def pb(text, size=9, align=0, color=colors.black):
    return p(text, FONT_BOLD, size, align, color)

# ══════════════════════════════════════════════════════════════
#  1. PHIEU LUONG CA NHAN (PDF) - ĐÃ FIX KHỚP VỚI EXCEL
# ══════════════════════════════════════════════════════════════
def generate(data, output_path):
    doc = SimpleDocTemplate(output_path, pagesize=A4, leftMargin=15*mm, rightMargin=15*mm, topMargin=15*mm, bottomMargin=15*mm)
    W = A4[0] - 30*mm

    # Thuc nhan = Tong thu nhap (cot trai) - Tong khoan tru NLD + phat + ung - bu thuong khac.
    # Khong lay finalSalary tu JSON: ban ghi SalaryHistory / backend co the lech so voi bang in (vd. snapshot cu).
    base = data.get('baseSalary', 0) or 0
    cc = data.get('chuyenCan', 0) or 0
    at = data.get('anTrua', 0) or 0
    htk = data.get('hoTroKhac', 0) or 0
    hh = data.get('hoaHong', 0) or 0
    tk = data.get('thuongKhac', 0) or 0
    ins = data.get('insuranceSalary', base) or base

    bhxh_cty = data.get('bhxhCty', round(ins * 0.175))
    bhyt_cty = data.get('bhytCty', round(ins * 0.03))
    bhtn_cty = data.get('bhtnCty', round(ins * 0.01))
    total_bh_cty = bhxh_cty + bhyt_cty + bhtn_cty

    bhxh = data.get('bhxhNld', round(ins * 0.08))
    bhyt = data.get('bhytNld', round(ins * 0.015))
    bhtn = data.get('bhtnNld', round(ins * 0.01))
    total_bh = bhxh + bhyt + bhtn

    half = data.get('halfDayDeduction', 0) or 0
    att_fine = data.get('attendanceFines', 0) or 0
    full_day = data.get('fullDayAbsenceDeduction', 0) or 0
    mf = data.get('manualFines', 0) or 0
    tu = data.get('tamUng', 0) or 0

    total_in = base + cc + at + htk + hh
    total_out = total_bh + half + att_fine + full_day + mf + tu - tk
    net_slip = int(round(total_in - total_out))

    story = []

    h = [[pb("Cong ty TNHH Fly Visa"), '', p("Luong thuc te"), pb(fmt(net_slip), align=2)],
         [p("MST: 0316444315"), '', p("Ngay cong di lam"), pb(str(data.get('workDays', 0)), align=2)],
         [p("DC: 219A Duong No Trang Long, Phuong Binh Thanh, TP.HCM"), '', '', '']]
    ht = Table(h, colWidths=[W*0.42, W*0.1, W*0.26, W*0.22])
    ht.setStyle(TableStyle([('BOX', (0, 0), (-1, -1), 0.5, colors.black), ('GRID', (0, 0), (-1, -1), 0.3, colors.grey), ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'), ('TOPPADDING', (0, 0), (-1, -1), 3), ('BOTTOMPADDING', (0, 0), (-1, -1), 3), ('LEFTPADDING', (0, 0), (-1, -1), 5)]))
    story.append(ht)
    story.append(Spacer(1, 4*mm))

    story.append(pb("PHIEU LUONG", size=12, align=1))
    story.append(Spacer(1, 1*mm))
    parts = str(data.get('monthYear', '')).split('/')
    month_text = f"THANG {parts[0]} NAM {parts[1]}" if len(parts) == 2 else data.get('monthYear', '')
    story.append(pb(month_text, size=11, align=1))
    story.append(Spacer(1, 3*mm))

    info = [[pb("Ma Nhan Vien"), p(data.get('employeeCode', '')), pb("Luong thuc te"), pb(fmt(net_slip), align=2)],
            [pb("Ho Va Ten"), pb(data.get('name', '')), pb("Ngay cong di lam"), pb(str(data.get('workDays', 0)), align=2)],
            [pb("Chuc Danh"), p(data.get('role', '')), '', '']]
    it = Table(info, colWidths=[W*0.2, W*0.3, W*0.25, W*0.25])
    it.setStyle(TableStyle([('BOX', (0, 0), (-1, -1), 0.5, colors.black), ('GRID', (0, 0), (-1, -1), 0.3, colors.grey), ('SPAN', (1, 2), (3, 2)), ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'), ('TOPPADDING', (0, 0), (-1, -1), 4), ('BOTTOMPADDING', (0, 0), (-1, -1), 4), ('LEFTPADDING', (0, 0), (-1, -1), 5)]))
    story.append(it)
    work_dates = data.get('workDates') or []
    if work_dates:
        ds = ', '.join(str(x) for x in work_dates)
        if len(ds) > 500:
            ds = ds[:497] + '...'
        story.append(Spacer(1, 2*mm))
        story.append(pb("Chi tiet ngay di lam trong thang:", size=8, align=0))
        story.append(p(ds, size=7, align=0))
    story.append(Spacer(1, 3*mm))

    def row(s1, l1, v1, s2='', l2='', v2=''):
        vv1 = fmt(v1) if isinstance(v1, (int, float)) and v1 else (str(v1) if v1 else '0')
        vv2 = fmt(v2) if isinstance(v2, (int, float)) and v2 else (str(v2) if v2 else '')
        return [p(s1), p(l1), p(vv1, align=2), p(s2), p(l2), p(vv2, align=2)]

    # ---- BẢNG RENDER (Khớp Excel 100%) ----
    rows = [
        [pb("STT", align=1), pb("Cac Khoan Thu Nhap"), p(''), pb("STT", align=1), pb("Cac Khoan Tru Vao Luong"), p('')],
        row('1', 'Luong co ban', base, '1', 'Chi phi Bao Hiem (DN)', '............'),
        row('2', 'Phu Cap:', '............', '1,1', 'BHXH (17,5%)', bhxh_cty),
        row('2,1', 'Chuyen can', cc, '1,2', 'BHYT (3%)', bhyt_cty),
        row('2,2', 'An trua', at, '1,3', 'BHTN (1%)', bhtn_cty),
        row('2,3', 'Ho tro khac', htk, '1,4', 'Tong Chi phi DN', total_bh_cty),
        row('2,4', 'Hoa hong / Thuong KPI', hh, '2', 'Trich vao Luong (NLD)', '............'),
        row('2,5', '', '', '2,5', 'Thuong khac (bu tru khoan tru)', -tk if tk else ''),
        row('', '', '', '2,1', 'BHXH (8%)', bhxh),
        row('', '', '', '2,2', 'BHYT (1,5%)', bhyt),
        row('', '', '', '2,3', 'BHTN (1%)', bhtn),
        row('', 'Tong Thu Nhap', total_in, '2,4', 'Tong Trich vao', total_bh),
        row('', '', '', '3', 'Vang ca ngay', full_day),
        row('', '', '', '3,1', 'Vang 1/2 ngay', half),
        row('', '', '', '3,2', 'Phat di tre', att_fine),
        row('', '', '', '3,3', 'Phat khac', mf),
        row('', '', '', '4', 'Tam ung', tu),
        [pb("Tong Cong"), p(''), pb(fmt(total_in), align=2), pb("Tong Cong"), p(''), pb(fmt(total_out), align=2)],
    ]

    mt = Table(rows, colWidths=[W*0.06, W*0.26, W*0.15, W*0.06, W*0.32, W*0.15])
    mt.setStyle(TableStyle([
        ('BOX', (0, 0), (-1, -1), 0.5, colors.black), 
        ('GRID', (0, 0), (-1, -1), 0.3, colors.grey), 
        ('BACKGROUND', (0, 0), (-1, 0), colors.whitesmoke), 
        ('BACKGROUND', (0, -1), (-1, -1), colors.whitesmoke), 
        ('LINEAFTER', (2, 0), (2, -1), 1, colors.black), 
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'), 
        ('TOPPADDING', (0, 0), (-1, -1), 3), 
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3), 
        ('LEFTPADDING', (0, 0), (-1, -1), 4)
    ]))
    story.append(mt)
    story.append(Spacer(1, 3*mm))

    ft = Table([[pb("Tong So Tien Luong Thuc Nhan", size=10), p(''), p(''), pb(fmt(net_slip), size=12, align=2, color=colors.HexColor('#1a56db'))]], colWidths=[W*0.45, W*0.1, W*0.1, W*0.35])
    ft.setStyle(TableStyle([('BOX', (0, 0), (-1, -1), 1.5, colors.HexColor('#1a56db')), ('SPAN', (0, 0), (2, 0)), ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'), ('TOPPADDING', (0, 0), (-1, -1), 7), ('BOTTOMPADDING', (0, 0), (-1, -1), 7), ('LEFTPADDING', (0, 0), (-1, -1), 6)]))
    story.append(ft)
    story.append(Spacer(1, 5*mm))

    sig = Table([[pb("Nguoi lap phieu", align=1), p(''), pb("Nguoi nhan tien", align=1)], [p("(Ky va ghi ro ho ten)", align=1), p(''), p("(Ky va ghi ro ho ten)", align=1)], [p('')]*3, [p('')]*3, [p('')]*3], colWidths=[W*0.4, W*0.2, W*0.4])
    sig.setStyle(TableStyle([('VALIGN', (0, 0), (-1, -1), 'MIDDLE')]))
    story.append(sig)
    doc.build(story)

# ══════════════════════════════════════════════════════════════
#  2. BANG LUONG TONG (PDF)
# ══════════════════════════════════════════════════════════════
def generate_summary(data, output_path):
    doc = SimpleDocTemplate(output_path, pagesize=landscape(A3), leftMargin=10*mm, rightMargin=10*mm, topMargin=10*mm, bottomMargin=10*mm)
    W = A3[1] - 20*mm
    story = []

    employees = data.get('employees', [])
    parts = str(data.get('monthYear', '')).split('/')
    month_text = f"THANG {parts[0]} NAM {parts[1]}" if len(parts) == 2 else data.get('monthYear', '')

    ht = Table([[pb("CONG TY TNHH FLY VISA", size=9), '', '', ''], [p("MST: 0316444315", size=8), '', '', ''], [p("DC: 219A Duong No Trang Long, Phuong Binh Thanh, TP.HCM", size=8), '', '', '']], colWidths=[W*0.5, W*0.2, W*0.15, W*0.15])
    ht.setStyle(TableStyle([('SPAN', (0, 0), (3, 0)), ('SPAN', (0, 1), (3, 1)), ('SPAN', (0, 2), (3, 2)), ('TOPPADDING', (0, 0), (-1, -1), 2), ('BOTTOMPADDING', (0, 0), (-1, -1), 2), ('LEFTPADDING', (0, 0), (-1, -1), 0)]))
    story.append(ht)
    story.append(Spacer(1, 3*mm))

    story.append(pb("BANG LUONG NHAN VIEN", size=13, align=1))
    story.append(Spacer(1, 1*mm))
    story.append(pb(month_text, size=11, align=1))
    story.append(Spacer(1, 4*mm))

    fs = 7
    def cp(text, size=None, bold=False, align=1): return pb(text, size=size or fs, align=align) if bold else p(text, size=size or fs, align=align)

    # Khau tru: T vang ca ngay; U gop phat di tre + tam ung (theo yeu cau hien thi cot U)
    header1 = [cp("STT", bold=True), cp("Ho va ten", bold=True), cp("Chuc vu", bold=True), cp("Luong co ban", bold=True), cp("Phu cap", bold=True), cp(""), cp(""), cp(""), cp(""), cp("Tong thu nhap", bold=True), cp("Luong dong BH", bold=True), cp("Cac khoan trich chi phi DN", bold=True), cp(""), cp(""), cp(""), cp("Cac khoan trich vao luong", bold=True), cp(""), cp(""), cp(""), cp("Vang\nca ngay", bold=True), cp("Phat di tre +\nTam ung", bold=True), cp("Tru\nnua ngay", bold=True), cp("Phat\nkhac", bold=True), cp("Thuong khac\n(bu tru)", bold=True), cp("Tong tru\n(rong)", bold=True), cp("Thuc linh", bold=True), cp("Ngay cong", bold=True)]
    header2 = [cp(""), cp(""), cp(""), cp(""), cp("Chuyen can", bold=True), cp("An trua", bold=True), cp("Ho tro khac", bold=True), cp("Hoa hong", bold=True), cp("-", bold=True), cp(""), cp(""), cp("BHXH\n(17,5%)", bold=True), cp("BHYT\n(3%)", bold=True), cp("BHTN\n(1%)", bold=True), cp("Tong", bold=True), cp("BHXH\n(8%)", bold=True), cp("BHYT\n(1,5%)", bold=True), cp("BHTN\n(1%)", bold=True), cp("Tong", bold=True), cp(""), cp(""), cp(""), cp(""), cp(""), cp(""), cp(""), cp("")]
    # 27 cot (A–AA): T vang, U phat tre+tam ung, V–X ..., Y tong tru, Z thuc linh, AA ngay cong
    cws = [W*0.05, W*0.08, W*0.04, W*0.046, W*0.029, W*0.027, W*0.027, W*0.029, W*0.029, W*0.044, W*0.038, W*0.029, W*0.023, W*0.023, W*0.029, W*0.025, W*0.021, W*0.021, W*0.027, W*0.03, W*0.035, W*0.03, W*0.03, W*0.032, W*0.036, W*0.058, W*0.078]
    table_rows = [header1, header2]

    for i, emp in enumerate(employees, start=1):
        def cv(v): return cp(fmt(v) if v else '0', align=2)
        hh = emp.get('hoaHong', 0) or 0
        cc = emp.get('chuyenCan', 0) or 0
        at = emp.get('anTrua', 0) or 0
        htk = emp.get('hoTroKhac', 0) or 0
        base = emp.get('baseSalary', 0) or 0
        thk = emp.get('thuongKhac', 0) or 0
        full_a = emp.get('fullDayAbsenceDeduction', 0) or 0
        half_d = emp.get('halfDayDeduction', 0) or 0
        att_f = emp.get('attendanceFines', 0) or 0
        mf = emp.get('manualFines', 0) or 0
        tu = emp.get('tamUng', 0) or 0
        att_plus_tam = att_f + tu
        tong_thu_nhap_khong_thuong_khac = base + cc + at + htk + hh
        total_tru_rong = full_a + att_plus_tam + half_d + mf - thk
        table_rows.append([
            cp(str(i), align=1), p(emp.get('name', ''), size=fs), p(emp.get('role', ''), size=fs, align=1),
            cv(base), cv(cc), cv(at), cv(htk), cv(hh), cp('-', align=2), cv(tong_thu_nhap_khong_thuong_khac), cv(emp.get('insuranceSalary')),
            cv(emp.get('bhxhCty')), cv(emp.get('bhytCty')), cv(emp.get('bhtnCty')), cv(emp.get('totalCty')), cv(emp.get('bhxhNld')), cv(emp.get('bhytNld')), cv(emp.get('bhtnNld')), cv(emp.get('totalNld')),
            cv(full_a), cv(att_plus_tam), cv(half_d), cv(mf), cv(thk), cv(total_tru_rong), cv(emp.get('finalSalary')), cp(str(emp.get('workDays', 0)), align=1),
        ])

    def total_field(field): return sum(e.get(field, 0) or 0 for e in employees)
    tot_base, tot_cc, tot_at, tot_htk, tot_hh = total_field('baseSalary'), total_field('chuyenCan'), total_field('anTrua'), total_field('hoTroKhac'), total_field('hoaHong')
    tot_thk = total_field('thuongKhac')
    tot_full = total_field('fullDayAbsenceDeduction')
    tot_half = total_field('halfDayDeduction')
    tot_att = total_field('attendanceFines')
    tot_mf = total_field('manualFines')
    tot_tam = total_field('tamUng')
    tot_tong_thu = tot_base + tot_cc + tot_at + tot_htk + tot_hh
    tot_att_tam = tot_att + tot_tam
    tot_tru_rong = tot_full + tot_att_tam + tot_half + tot_mf - tot_thk

    def tv(v): return cp(fmt(v) if v else '0', bold=True, align=2)
    table_rows.append([cp("TONG", bold=True, align=1), cp(""), cp(""), tv(tot_base), tv(tot_cc), tv(tot_at), tv(tot_htk), tv(tot_hh), cp('-', bold=True, align=2), tv(tot_tong_thu), tv(total_field('insuranceSalary')), tv(total_field('bhxhCty')), tv(total_field('bhytCty')), tv(total_field('bhtnCty')), tv(total_field('totalCty')), tv(total_field('bhxhNld')), tv(total_field('bhytNld')), tv(total_field('bhtnNld')), tv(total_field('totalNld')), tv(tot_full), tv(tot_att_tam), tv(tot_half), tv(tot_mf), tv(tot_thk), tv(tot_tru_rong), tv(total_field('finalSalary')), cp(str(int(total_field('workDays'))), align=1, bold=True)])

    tbl = Table(table_rows, colWidths=cws, repeatRows=2)
    n_data = len(table_rows)
    tbl.setStyle(TableStyle([
        ('BOX', (0, 0), (-1, -1), 1, colors.black), ('GRID', (0, 0), (-1, -1), 0.3, colors.grey),
        ('SPAN', (0, 0), (0, 1)), ('SPAN', (1, 0), (1, 1)), ('SPAN', (2, 0), (2, 1)), ('SPAN', (3, 0), (3, 1)), ('SPAN', (4, 0), (8, 0)), ('SPAN', (9, 0), (9, 1)), ('SPAN', (10, 0), (10, 1)), ('SPAN', (11, 0), (14, 0)), ('SPAN', (15, 0), (18, 0)),
        ('SPAN', (19, 0), (19, 1)), ('SPAN', (20, 0), (20, 1)), ('SPAN', (21, 0), (21, 1)), ('SPAN', (22, 0), (22, 1)), ('SPAN', (23, 0), (23, 1)), ('SPAN', (24, 0), (24, 1)), ('SPAN', (25, 0), (25, 1)), ('SPAN', (26, 0), (26, 1)),
        ('SPAN', (0, n_data - 1), (2, n_data - 1)),
        ('BACKGROUND', (0, 0), (-1, 1), colors.HexColor('#FFA500')), ('BACKGROUND', (0, n_data - 1), (-1, n_data - 1), colors.HexColor('#FFA500')),
        *[('BACKGROUND', (0, r), (-1, r), colors.HexColor('#FFF3E0')) for r in range(2, n_data - 1) if (r - 2) % 2 == 1],
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'), ('ALIGN', (0, 0), (-1, -1), 'CENTER'), ('TOPPADDING', (0, 0), (-1, -1), 2), ('BOTTOMPADDING', (0, 0), (-1, -1), 2), ('LEFTPADDING', (0, 0), (-1, -1), 2), ('RIGHTPADDING', (0, 0), (-1, -1), 2),
        ('LINEAFTER', (8, 0), (8, -1), 1, colors.black), ('LINEAFTER', (14, 0), (14, -1), 1, colors.black), ('LINEAFTER', (18, 0), (18, -1), 1, colors.black),
    ]))
    story.append(tbl)
    story.append(Spacer(1, 8*mm))

    sig = Table([[pb("Nguoi lap bieu", align=1), p(''), pb("Ke toan", align=1), p(''), pb("Giam doc", align=1)], [p("(Ky, ghi ro ho ten)", size=8, align=1), p(''), p("(Ky, ghi ro ho ten)", size=8, align=1), p(''), p("(Ky, ghi ro ho ten)", size=8, align=1)], [p('')]*5, [p('')]*5, [p('')]*5], colWidths=[W*0.25, W*0.1, W*0.25, W*0.1, W*0.3])
    sig.setStyle(TableStyle([('TOPPADDING', (0, 0), (-1, -1), 4), ('BOTTOMPADDING', (0, 0), (-1, -1), 4), ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')]))
    story.append(sig)
    doc.build(story)

# ══════════════════════════════════════════════════════════════
#  CÁC HÀM HỖ TRỢ EXCEL (openpyxl)
# ══════════════════════════════════════════════════════════════
def get_bd():
    from openpyxl.styles import Border, Side
    return Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

def style_merge(ws, cell_range, value='', bold=False, size=10, h='left', v='center', fill=None, color='000000', num_fmt=None):
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import range_boundaries
    ws.merge_cells(cell_range)
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    fnt = Font(name='Arial', bold=bold, size=size, color=color)
    alg = Alignment(horizontal=h, vertical=v, wrap_text=True)
    fil = PatternFill("solid", fgColor=fill) if fill else None
    border = get_bd()
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            if fil: cell.fill = fil
    top_cell = ws.cell(row=min_row, column=min_col)
    top_cell.value = value
    top_cell.font = fnt
    top_cell.alignment = alg
    if num_fmt: top_cell.number_format = num_fmt
    return top_cell

def sc(ws, cell_ref, value='', bold=False, size=10, h='left', v='center', fill=None, num_fmt=None, color='000000', bdr=False, wrap=False):
    from openpyxl.styles import Font, PatternFill, Alignment
    c = ws[cell_ref]
    c.value = value
    c.font = Font(name='Arial', bold=bold, size=size, color=color)
    c.alignment = Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    if fill: c.fill = PatternFill("solid", fgColor=fill)
    if bdr: c.border = get_bd()
    if num_fmt: c.number_format = num_fmt
    return c

def _add_history_sheet(wb, employees, month_text):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    ws = wb.create_sheet("Chi tiet lich su")
    ORANGE, LGRAY, RED_LIGHT, GREEN_LIGHT = "FFA500", "D3D3D3", "FFE6E6", "E6F4E6"
    r = 1
    for col, name in zip('ABCDEFGH', ['STT', 'Ma NV', 'Ho va ten', 'Ngay', 'Loai giao dich', 'Kiểu', 'Chi tiet', 'So tien']):
        sc(ws, f'{col}{r}', name, bold=True, fill=ORANGE, bdr=True)
    ws.row_dimensions[r].height = 16
    row, stt = 2, 0
    for emp in employees:
        emp_code, emp_name = emp.get('employeeCode', ''), emp.get('name', '')
        hh = emp.get('hoaHong', 0) or 0
        if hh > 0:
            stt += 1
            for col, val, align in zip('ABCDEFGH', [stt, emp_code, emp_name, '', 'Hoa hồng / Thưởng KPI', 'CỘNG', 'Doanh số & thưởng HH', hh], ['center','center','left','center','left','center','left','right']):
                sc(ws, f'{col}{row}', val, size=8, h=align, fill=GREEN_LIGHT, bold=(col in 'EF'), bdr=True)
            ws[f'H{row}'].number_format = '#,##0'
            ws.row_dimensions[row].height = 12; row += 1
        thk = emp.get('thuongKhac', 0) or 0
        if thk > 0:
            stt += 1
            for col, val, align in zip('ABCDEFGH', [stt, emp_code, emp_name, '', 'Thưởng khác', 'CỘNG', 'Thưởng không gắn KPI (điều chỉnh)', thk], ['center','center','left','center','left','center','left','right']):
                sc(ws, f'{col}{row}', val, size=8, h=align, fill=GREEN_LIGHT, bold=(col in 'EF'), bdr=True)
            ws[f'H{row}'].number_format = '#,##0'
            ws.row_dimensions[row].height = 12; row += 1
        for record in emp.get('absenceRecords', []):
            stt += 1
            for col, val, align in zip('ABCDEFGH', [stt, emp_code, emp_name, record.get('date', ''), 'Vắng cả ngày', 'TRỪ', record.get('status', ''), record.get('amount', 0)], ['center','center','left','center','left','center','left','right']):
                sc(ws, f'{col}{row}', val, size=8, h=align, fill=RED_LIGHT, bold=(col == 'F'), bdr=True)
            ws[f'H{row}'].number_format = '#,##0'; row += 1
        for record in emp.get('fineRecords', []):
            stt += 1
            for col, val, align in zip('ABCDEFGH', [stt, emp_code, emp_name, record.get('date', ''), 'Phạt đi trễ', 'TRỪ', record.get('status', ''), record.get('amount', 0)], ['center','center','left','center','left','center','left','right']):
                sc(ws, f'{col}{row}', val, size=8, h=align, bold=(col == 'F'), bdr=True)
            ws[f'H{row}'].number_format = '#,##0'; row += 1
        for record in emp.get('lateRecords', []):
            stt += 1
            for col, val, align in zip('ABCDEFGH', [stt, emp_code, emp_name, record.get('date', ''), 'Nửa ngày công', 'TRỪ', record.get('status', ''), record.get('amount', 0)], ['center','center','left','center','left','center','left','right']):
                sc(ws, f'{col}{row}', val, size=8, h=align, bold=(col == 'F'), bdr=True)
            ws[f'H{row}'].number_format = '#,##0'; row += 1
        mf = emp.get('manualFines', 0) or 0
        if mf > 0:
            stt += 1
            for col, val, align in zip('ABCDEFGH', [stt, emp_code, emp_name, '', 'Phạt khác', 'TRỪ', 'Phạt theo quy định', mf], ['center','center','left','center','left','center','left','right']):
                sc(ws, f'{col}{row}', val, size=8, h=align, bold=(col in 'EF'), bdr=True)
            ws[f'H{row}'].number_format = '#,##0'; row += 1
        for record in emp.get('advanceRecords', []):
            stt += 1
            for col, val, align in zip('ABCDEFGH', [stt, emp_code, emp_name, record.get('date', ''), 'Ứng lương', 'TRỪ', record.get('note', ''), record.get('amount', 0)], ['center','center','left','center','left','center','left','right']):
                sc(ws, f'{col}{row}', val, size=8, h=align, bold=(col == 'F'), bdr=True)
            ws[f'H{row}'].number_format = '#,##0'; row += 1
            
    ws.column_dimensions['A'].width, ws.column_dimensions['B'].width, ws.column_dimensions['C'].width, ws.column_dimensions['D'].width = 5, 10, 20, 12
    ws.column_dimensions['E'].width, ws.column_dimensions['F'].width, ws.column_dimensions['G'].width, ws.column_dimensions['H'].width = 18, 8, 30, 12

def _add_slip_sheets(wb, employees, month_text):
    from openpyxl.styles import Border, Side
    ORANGE, BLUE, LGRAY = "FFA500", "1a56db", "F5F5F5"

    for emp in employees:
        sheet_name = emp.get('name', 'NV').strip()[:28]
        for ch in r'\/*?[]': sheet_name = sheet_name.replace(ch, '')
        ws = wb.create_sheet(title=sheet_name)

        base = emp.get('baseSalary', 0) or 0
        cc = emp.get('chuyenCan', 0) or 0
        at = emp.get('anTrua', 0) or 0
        htk = emp.get('hoTroKhac', 0) or 0
        hh = emp.get('hoaHong', 0) or 0
        tk = emp.get('thuongKhac', 0) or 0
        ins = emp.get('insuranceSalary', base) or base
        
        bhxh_cty = emp.get('bhxhCty', 0) or 0
        bhyt_cty = emp.get('bhytCty', 0) or 0
        bhtn_cty = emp.get('bhtnCty', 0) or 0
        total_bh_cty = bhxh_cty + bhyt_cty + bhtn_cty
        
        bhxh = emp.get('bhxhNld', round(ins * 0.08))
        bhyt = emp.get('bhytNld', round(ins * 0.015))
        bhtn = emp.get('bhtnNld', round(ins * 0.01))
        
        full_day_abs = emp.get('fullDayAbsenceDeduction', 0) or 0
        half = emp.get('halfDayDeduction', 0) or 0
        att_fine = emp.get('attendanceFines', 0) or 0
        manual_fine = emp.get('manualFines', 0) or 0
        tu = emp.get('tamUng', 0) or 0
        wdays = emp.get('workDays', 0) or 0
        
        total_in = base + cc + at + htk + hh
        total_bh = bhxh + bhyt + bhtn
        total_out = total_bh + full_day_abs + half + tu + att_fine + manual_fine - tk
        final = int(round(total_in - total_out))

        style_merge(ws, 'A1:G1', 'Cong ty TNHH Fly Visa', bold=True, size=11)
        style_merge(ws, 'A2:G2', 'MST: 0316444315', size=9)
        style_merge(ws, 'A3:G3', 'DC: 219A Duong No Trang Long, Phuong Binh Thanh, TP.HCM', size=9)
        ws.row_dimensions[4].height = 6

        style_merge(ws, 'A5:G5', 'PHIEU LUONG', bold=True, size=14, h='center')
        style_merge(ws, 'A6:G6', month_text, bold=True, size=12, h='center')
        for r in range(1, 7):
            for c in range(1, 8): ws.cell(row=r, column=c).border = None

        ws.row_dimensions[5].height, ws.row_dimensions[6].height = 20, 18
        ws.row_dimensions[7].height = 4

        sc(ws, 'A8', 'Ma Nhan Vien', bold=True, size=9, bdr=True, fill=LGRAY)
        style_merge(ws, 'B8:D8', emp.get('employeeCode', ''), size=9)
        sc(ws, 'E8', 'Luong thuc te', bold=True, size=9, bdr=True, fill=LGRAY)
        style_merge(ws, 'F8:G8', final, size=10, bold=True, h='right', num_fmt='#,##0')

        sc(ws, 'A9', 'Ho Va Ten', bold=True, size=9, bdr=True, fill=LGRAY)
        style_merge(ws, 'B9:D9', emp.get('name', ''), size=9, bold=True)
        sc(ws, 'E9', 'Ngay cong di lam', bold=True, size=9, bdr=True, fill=LGRAY)
        style_merge(ws, 'F9:G9', wdays, size=10, bold=True, h='right')

        sc(ws, 'A10', 'Chuc Danh', bold=True, size=9, bdr=True, fill=LGRAY)
        style_merge(ws, 'B10:G10', emp.get('role', ''), size=9)
        for r in (8, 9, 10): ws.row_dimensions[r].height = 16

        current_row = 11
        wdates_slip = emp.get('workDates') or []
        if wdates_slip:
            ds_slip = ', '.join(str(x) for x in wdates_slip)
            if len(ds_slip) > 450:
                ds_slip = ds_slip[:447] + '...'
            style_merge(ws, f'A{current_row}:G{current_row}', f'Chi tiet ngay di lam: {ds_slip}', size=8)
            ws.row_dimensions[current_row].height = 28
            current_row += 1

        def render_detail_group(title, fill_color, records, fmt_str):
            nonlocal current_row
            if not records: return
            ws.row_dimensions[current_row].height = 4; current_row += 1
            style_merge(ws, f'A{current_row}:G{current_row}', title, bold=True, size=8, fill=fill_color)
            ws.row_dimensions[current_row].height = 14; current_row += 1
            for rec in records:
                style_merge(ws, f'A{current_row}:G{current_row}', fmt_str(rec), size=8)
                ws.row_dimensions[current_row].height = 12; current_row += 1

        render_detail_group('Chi tiết các ngày vắng (cả ngày)', 'FFE6E6', emp.get('absenceRecords', []), lambda r: f"{r.get('date', '')} - Vắng - Trừ: {r.get('amount', 0):,.0f}đ")
        render_detail_group('Chi tiết phạt đi trễ', 'FFF4E6', emp.get('fineRecords', []), lambda r: f"{r.get('date', '')} - Phạt: {r.get('amount', 0):,.0f}đ")
        render_detail_group('Chi tiết các ngày đi trễ / nửa ngày công', 'FFF3E0', emp.get('lateRecords', []), lambda r: f"{r.get('date', '')} - {r.get('status', '')} - Trừ: {r.get('amount', 0):,.0f}đ")
        render_detail_group('Chi tiết các lần ứng lương', 'FFF3E0', emp.get('advanceRecords', []), lambda r: f"{r.get('date', '')} - Ứng: {r.get('amount', 0):,.0f}đ - {r.get('note', '')}")

        row_header = current_row + 2
        ws.row_dimensions[current_row].height = 4; ws.row_dimensions[current_row + 1].height = 4

        sc(ws, f'A{row_header}', 'STT', bold=True, size=8, h='center', fill=ORANGE, bdr=True)
        sc(ws, f'B{row_header}', 'Cac Khoan Thu Nhap', bold=True, size=8, h='center', fill=ORANGE, bdr=True)
        sc(ws, f'C{row_header}', '', bold=True, size=8, fill=ORANGE, bdr=True)
        sc(ws, f'D{row_header}', 'STT', bold=True, size=8, h='center', fill=ORANGE, bdr=True)
        style_merge(ws, f'E{row_header}:F{row_header}', 'Cac Khoan Tru Vao Luong', bold=True, size=8, h='center', fill=ORANGE)
        sc(ws, f'G{row_header}', 'So tien', bold=True, size=8, h='center', fill=ORANGE, bdr=True)
        ws.row_dimensions[row_header].height = 16

        def dr(row, s1, l1, v1, s2, l2, v2):
            sc(ws, f'A{row}', s1, size=8, h='center', bdr=True)
            sc(ws, f'B{row}', l1, size=8, bdr=True)
            sc(ws, f'C{row}', v1 if v1 else (0 if isinstance(v1, (int,float)) else ''), size=8, h='right', num_fmt='#,##0' if isinstance(v1, (int,float)) else None, bdr=True)
            sc(ws, f'D{row}', s2, size=8, h='center', bdr=True)
            style_merge(ws, f'E{row}:F{row}', l2, size=8)
            sc(ws, f'G{row}', v2 if v2 else (0 if isinstance(v2, (int,float)) else ''), size=8, h='right', num_fmt='#,##0' if isinstance(v2, (int,float)) else None, bdr=True)
            ws.row_dimensions[row].height = 15

        start_data_row = row_header + 1
        dr(start_data_row,     '1',   'Luong co ban',        base, '1',   'Chi phi Bao Hiem (DN)',        '............')
        dr(start_data_row + 1, '2',   'Phu Cap:',            '............', '1,1', 'BHXH (17,5%)',          bhxh_cty)
        dr(start_data_row + 2, '2,1', 'Chuyen can',          cc,   '1,2', 'BHYT (3%)',                   bhyt_cty)
        dr(start_data_row + 3, '2,2', 'An trua',             at,   '1,3', 'BHTN (1%)',                   bhtn_cty)
        dr(start_data_row + 4, '2,3', 'Ho tro khac',         htk,  '1,4', 'Tong Chi phi DN',             total_bh_cty)
        dr(start_data_row + 5, '2,4', 'Hoa hong / Thuong KPI', hh, '2',   'Trich vao Luong (NLD)',       '............')
        dr(start_data_row + 6, '2,5', '',                       '',   '2,5', 'Thuong khac (bu tru khoan tru)', -tk if tk else 0)
        dr(start_data_row + 7, '',    '',                       '',   '2,1', 'BHXH (8%)',                   bhxh)
        dr(start_data_row + 8, '',    '',                       '',   '2,2', 'BHYT (1,5%)',                 bhyt)
        dr(start_data_row + 9, '',    '',                       '',   '2,3', 'BHTN (1%)',                   bhtn)
        dr(start_data_row + 10, '',   'Tong Thu Nhap',          total_in, '2,4', 'Tong Trich vao',          total_bh)
        dr(start_data_row + 11, '',   '',                       '',   '3', 'Vang ca ngay',                  full_day_abs)
        dr(start_data_row + 12, '',   '',                       '',   '3,1', 'Vang 1/2 ngay',               half)
        dr(start_data_row + 13, '',   '',                       '',   '3,2', 'Phat di tre',                 att_fine)
        dr(start_data_row + 14, '',   '',                       '',   '3,3', 'Phat khac',                   manual_fine)
        dr(start_data_row + 15, '',   '',                       '',   '4',   'Tam ung',                     tu)

        row_dong_tong = start_data_row + 16
        style_merge(ws, f'A{row_dong_tong}:B{row_dong_tong}', 'Tong Cong', bold=True, size=9, h='center', fill=ORANGE)
        sc(ws, f'C{row_dong_tong}', total_in, bold=True, size=9, h='right', fill=ORANGE, bdr=True, num_fmt='#,##0')
        style_merge(ws, f'D{row_dong_tong}:F{row_dong_tong}', 'Tong Cong', bold=True, size=9, h='center', fill=ORANGE)
        sc(ws, f'G{row_dong_tong}', total_out, bold=True, size=9, h='right', fill=ORANGE, bdr=True, num_fmt='#,##0')
        ws.row_dimensions[row_dong_tong].height = 16

        row_tong_thuc = row_dong_tong + 2
        ws.row_dimensions[row_tong_thuc - 1].height = 6
        thucnhan_bd = Border(left=Side(style='medium', color=BLUE), right=Side(style='medium', color=BLUE), top=Side(style='medium', color=BLUE), bottom=Side(style='medium', color=BLUE))
        style_merge(ws, f'A{row_tong_thuc}:E{row_tong_thuc}', 'Tong So Tien Luong Thuc Nhan', bold=True, size=11)
        style_merge(ws, f'F{row_tong_thuc}:G{row_tong_thuc}', final, bold=True, size=12, h='right', color=BLUE, num_fmt='#,##0')
        for c in range(1, 8):
            ws.cell(row=row_tong_thuc, column=c).border = Border(top=thucnhan_bd.top, bottom=thucnhan_bd.bottom, left=thucnhan_bd.left if c == 1 else None, right=thucnhan_bd.right if c == 7 else None)
        ws.row_dimensions[row_tong_thuc].height = 20

        row_bang_chu = row_tong_thuc + 2
        ws.row_dimensions[row_tong_thuc + 1].height = 6
        sc(ws, f'A{row_bang_chu}', 'Bang chu:', bold=True, size=9, bdr=True)
        style_merge(ws, f'B{row_bang_chu}:G{row_bang_chu}', '', size=9)
        ws.row_dimensions[row_bang_chu].height = 20

        # Đã cập nhật bề rộng cột A lên 8 để đẹp hơn
        ws.column_dimensions['A'].width, ws.column_dimensions['B'].width, ws.column_dimensions['C'].width, ws.column_dimensions['D'].width = 8, 22, 14, 6
        ws.column_dimensions['E'].width, ws.column_dimensions['F'].width, ws.column_dimensions['G'].width = 22, 4, 14
        ws.sheet_view.showGridLines = False

# ══════════════════════════════════════════════════════════════
#  3. BANG LUONG TONG (EXCEL)
# ══════════════════════════════════════════════════════════════
def generate_summary_excel(data, output_path):
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise Exception("Thieu thu vien: pip install openpyxl")

    employees = data.get('employees', [])
    parts = str(data.get('monthYear', '')).split('/')
    month_text = f"THANG {parts[0]} NAM {parts[1]}" if len(parts) == 2 else data.get('monthYear', '')

    wb = Workbook()
    ws = wb.active
    ws.title = "Bang Luong"

    ORANGE, LT_ORANGE, WHITE = "FFA500", "FFF3E0", "FFFFFF"
    NUM = '#,##0'
    LAST_COL_LETTER = 'AA'

    style_merge(ws, f'A1:{LAST_COL_LETTER}1', 'CONG TY TNHH FLY VISA', bold=True, size=11)
    style_merge(ws, f'A2:{LAST_COL_LETTER}2', 'MST: 0316444315', size=9)
    style_merge(ws, f'A3:{LAST_COL_LETTER}3', 'DC: 219A Duong No Trang Long, Phuong Binh Thanh, TP.HCM', size=9)
    style_merge(ws, f'A5:{LAST_COL_LETTER}5', 'BANG LUONG NHAN VIEN', bold=True, size=14, h='center')
    style_merge(ws, f'A6:{LAST_COL_LETTER}6', month_text, bold=True, size=12, h='center')

    for r in range(1, 7):
        for c in range(1, 28): ws.cell(row=r, column=c).border = None
    ws.row_dimensions[5].height, ws.row_dimensions[6].height = 22, 18

    HR1, HR2, DS = 8, 9, 10

    style_merge(ws, f'A{HR1}:A{HR2}', 'STT', bold=True, size=8, h='center', fill=ORANGE)
    style_merge(ws, f'B{HR1}:B{HR2}', 'Ho va ten', bold=True, size=8, h='center', fill=ORANGE)
    style_merge(ws, f'C{HR1}:C{HR2}', 'Chuc vu', bold=True, size=8, h='center', fill=ORANGE)
    style_merge(ws, f'D{HR1}:D{HR2}', 'Luong co ban', bold=True, size=8, h='center', fill=ORANGE)
    style_merge(ws, f'E{HR1}:I{HR1}', 'Phu cap', bold=True, size=8, h='center', fill=ORANGE)
    style_merge(ws, f'J{HR1}:J{HR2}', 'Tong thu nhap', bold=True, size=8, h='center', fill=ORANGE)
    style_merge(ws, f'K{HR1}:K{HR2}', 'Luong dong BH', bold=True, size=8, h='center', fill=ORANGE)
    style_merge(ws, f'L{HR1}:O{HR1}', 'Cac khoan trich chi phi DN', bold=True, size=8, h='center', fill=ORANGE)
    style_merge(ws, f'P{HR1}:S{HR1}', 'Cac khoan trich vao luong', bold=True, size=8, h='center', fill=ORANGE)
    style_merge(ws, f'T{HR1}:T{HR2}', 'Vang ca ngay\n(khong diem danh)', bold=True, size=8, h='center', fill=ORANGE)
    style_merge(ws, f'U{HR1}:U{HR2}', 'Phat di tre +\nTam ung', bold=True, size=8, h='center', fill=ORANGE)
    style_merge(ws, f'V{HR1}:V{HR2}', 'Tru nua ngay/\nve som/quen CO', bold=True, size=8, h='center', fill=ORANGE)
    style_merge(ws, f'W{HR1}:W{HR2}', 'Phat khac\n(CRM/thu cong)', bold=True, size=8, h='center', fill=ORANGE)
    style_merge(ws, f'X{HR1}:X{HR2}', 'Thuong khac\n(bu tru)', bold=True, size=8, h='center', fill=ORANGE)
    style_merge(ws, f'Y{HR1}:Y{HR2}', 'Tong tru\n(rong)', bold=True, size=8, h='center', fill=ORANGE)
    style_merge(ws, f'Z{HR1}:Z{HR2}', 'Thuc linh', bold=True, size=8, h='center', fill=ORANGE)
    style_merge(ws, f'AA{HR1}:AA{HR2}', 'Ngay cong', bold=True, size=8, h='center', fill=ORANGE)

    subs = {
        'E': 'Chuyen can', 'F': 'An trua', 'G': 'Ho tro khac', 'H': 'Hoa hong', 'I': '—',
        'L': 'BHXH\n(17,5%)', 'M': 'BHYT\n(3%)', 'N': 'BHTN\n(1%)', 'O': 'Tong',
        'P': 'BHXH\n(8%)', 'Q': 'BHYT\n(1,5%)', 'R': 'BHTN\n(1%)', 'S': 'Tong'
    }
    for col, text in subs.items():
        sc(ws, f'{col}{HR2}', text, bold=True, size=8, h='center', fill=ORANGE, bdr=True)

    ws.row_dimensions[HR1].height, ws.row_dimensions[HR2].height = 28, 28

    for i, emp in enumerate(employees, start=1):
        r = DS + i - 1
        fill_c = LT_ORANGE if i % 2 == 0 else WHITE

        base = money_amount(emp.get('baseSalary'))
        cc = money_amount(emp.get('chuyenCan'))
        at = money_amount(emp.get('anTrua'))
        htk = money_amount(emp.get('hoTroKhac'))
        hh = money_amount(emp.get('hoaHong'))
        thk = money_amount(emp.get('thuongKhac'))
        ngay = int(money_amount(emp.get('workDays')))
        tu = money_amount(emp.get('tamUng'))
        full_a = money_amount(emp.get('fullDayAbsenceDeduction'))
        half_d = money_amount(emp.get('halfDayDeduction'))
        att_f = money_amount(emp.get('attendanceFines'))
        mf = money_amount(emp.get('manualFines'))

        att_plus_tam = att_f + tu

        data_map = {
            'A': (i, 'center'), 'B': (emp.get('name', ''), 'left'), 'C': (emp.get('role', ''), 'left'),
            'D': (base, 'right'), 'E': (cc, 'right'), 'F': (at, 'right'), 'G': (htk, 'right'), 'H': (hh, 'right'), 'I': (0.0, 'right'),
            'T': (full_a, 'right'), 'U': (att_plus_tam, 'right'), 'V': (half_d, 'right'), 'W': (mf, 'right'), 'X': (thk, 'right'),
            'Z': (money_amount(emp.get('finalSalary')), 'right'), 'AA': (ngay, 'center')
        }

        for col, (val, align) in data_map.items():
            if col in ('B', 'C'):
                out_val = val
            elif col == 'A':
                out_val = i
            elif col == 'AA':
                out_val = ngay
            else:
                out_val = money_amount(val)
            c = sc(ws, f'{col}{r}', out_val, h=align, fill=fill_c, bdr=True)
            if col not in ('A', 'B', 'C', 'AA'):
                c.number_format = NUM

        formulas = {
            # D–H: lương CB + 3 phụ cấp + hoa hồng; I là cột “—” (0), không gộp thưởng khác (cột X).
            'J': f'=SUM(D{r}:H{r})', 'K': f'=D{r}',
            'L': f'=ROUND(K{r}*17.5%,0)', 'M': f'=ROUND(K{r}*3%,0)', 'N': f'=ROUND(K{r}*1%,0)', 'O': f'=SUM(L{r}:N{r})',
            'P': f'=ROUND(K{r}*8%,0)', 'Q': f'=ROUND(K{r}*1.5%,0)', 'R': f'=ROUND(K{r}*1%,0)', 'S': f'=SUM(P{r}:R{r})',
            'Y': f'=T{r}+U{r}+V{r}+W{r}-X{r}',
        }
        for col, f in formulas.items():
            c = sc(ws, f'{col}{r}', f, h='right', fill=fill_c, bdr=True)
            c.number_format = NUM
        ws.row_dimensions[r].height = 16

    TR = DS + len(employees)
    style_merge(ws, f'A{TR}:C{TR}', 'TONG CONG', bold=True, size=10, h='center', fill=ORANGE)
    for col_idx in range(4, 28):
        col = get_column_letter(col_idx)
        cell = sc(ws, f'{col}{TR}', f'=SUM({col}{DS}:{col}{TR-1})', bold=True, h='right', fill=ORANGE, bdr=True)
        cell.number_format = NUM
    ws.row_dimensions[TR].height = 18

    widths = {'A': 5, 'B': 22, 'C': 12, 'D': 14, 'E': 11, 'F': 10, 'G': 11, 'H': 11, 'I': 11, 'J': 13, 'K': 13, 'L': 11, 'M': 9, 'N': 9, 'O': 11, 'P': 9, 'Q': 9, 'R': 9, 'S': 11, 'T': 12, 'U': 13, 'V': 12, 'W': 11, 'X': 11, 'Y': 12, 'Z': 13, 'AA': 10}
    for c, w in widths.items(): ws.column_dimensions[c].width = w

    ws.freeze_panes = f'A{DS}'
    _add_history_sheet(wb, employees, month_text)
    _add_slip_sheets(wb, employees, month_text)
    
    wb.save(output_path)

# ══════════════════════════════════════════════════════════════
#  ENTRY POINT
# ══════════════════════════════════════════════════════════════
if __name__ == '__main__':
    if len(sys.argv) == 4 and sys.argv[1] == 'summary':
        with open(sys.argv[2], encoding='utf-8') as f: generate_summary(json.load(f), sys.argv[3])
    elif len(sys.argv) == 4 and sys.argv[1] == 'excel':
        with open(sys.argv[2], encoding='utf-8') as f: generate_summary_excel(json.load(f), sys.argv[3])
    elif len(sys.argv) == 4 and sys.argv[1] == 'slips':
        with open(sys.argv[2], encoding='utf-8') as f:
            try: from openpyxl import Workbook
            except ImportError: sys.exit(1)
            data = json.load(f)
            wb = Workbook()
            wb.remove(wb.active)
            parts = str(data.get('monthYear', '')).split('/')
            month_text = f"THANG {parts[0]} NAM {parts[1]}" if len(parts) == 2 else data.get('monthYear', '')
            _add_slip_sheets(wb, data.get('employees', []), month_text)
            wb.save(sys.argv[3])
    elif len(sys.argv) == 3:
        with open(sys.argv[1], encoding='utf-8') as f: generate(json.load(f), sys.argv[2])
    else:
        sys.exit(1)